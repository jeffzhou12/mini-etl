import openpyxl
import xml.etree.ElementTree as ET
from pathlib import Path
import base64

def excel_to_markdown(excel_path):
    """将 Excel 转换为 Markdown 表格"""
    wb = openpyxl.load_workbook(excel_path)
    markdown_output = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        markdown_output.append(f"\n### {sheet_name}\n")
        
        # 获取所有行数据
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            markdown_output.append("_（空表）_\n")
            continue
        
        # 过滤掉完全空白的行
        non_empty_rows = []
        for row in rows:
            if any(cell is not None and str(cell).strip() for cell in row):
                non_empty_rows.append(row)
        
        if not non_empty_rows:
            markdown_output.append("_（空表）_\n")
            continue
        
        # 获取最大列数
        max_cols = max(len(row) for row in non_empty_rows)
        
        # 表头
        header = non_empty_rows[0] if non_empty_rows else []
        header = [str(cell) if cell is not None else "" for cell in header]
        # 填充到最大列数
        header.extend([""] * (max_cols - len(header)))
        markdown_output.append("| " + " | ".join(header) + " |")
        
        # 分隔行
        markdown_output.append("| " + " | ".join(["---"] * max_cols) + " |")
        
        # 数据行
        for row in non_empty_rows[1:]:
            row_data = [str(cell) if cell is not None else "" for cell in row]
            row_data.extend([""] * (max_cols - len(row_data)))
            markdown_output.append("| " + " | ".join(row_data) + " |")
        
        markdown_output.append("")
    
    return "\n".join(markdown_output)

def export_drawio_info(drawio_path):
    """读取 Drawio 文件信息"""
    try:
        tree = ET.parse(drawio_path)
        root = tree.getroot()
        diagrams = root.findall('.//diagram')
        
        info = f"Draw.io 文件包含 {len(diagrams)} 个图表：\n"
        for diagram in diagrams:
            name = diagram.get('name', '未命名')
            info += f"- {name}\n"
        return info
    except Exception as e:
        return f"读取 Drawio 文件出错: {e}"

if __name__ == "__main__":
    # 转换 Excel
    excel_path = "etl_model_design.xlsx"
    if Path(excel_path).exists():
        print("正在转换 Excel 文件...")
        markdown = excel_to_markdown(excel_path)
        
        # 保存为独立文件
        with open("etl_model_design.md", "w", encoding="utf-8") as f:
            f.write("# ETL 模型设计\n")
            f.write(markdown)
        print("✓ Excel 已转换为 etl_model_design.md")
    else:
        print(f"✗ 未找到文件: {excel_path}")
    
    # 读取 Drawio 信息
    drawio_path = "etl_flow.drawio"
    if Path(drawio_path).exists():
        print("\n正在分析 Drawio 文件...")
        info = export_drawio_info(drawio_path)
        print(info)
        print("\n注意：Drawio 文件需要手动导出为 PNG/SVG")
        print("方法1: 使用 Draw.io Desktop 或在线版本 (https://app.diagrams.net/)")
        print("方法2: 使用 VS Code 的 Draw.io Integration 扩展")
    else:
        print(f"✗ 未找到文件: {drawio_path}")
